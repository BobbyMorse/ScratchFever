from openpyxl import load_workbook
wb = load_workbook('_mo_winners.xlsx', read_only=True, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'sheet: {sn}, rows={ws.max_row}')
ws = wb[wb.sheetnames[0]]
hits = 0
for row in ws.iter_rows(values_only=True):
    if row and any('GET IT N GO' in str(c).upper() for c in row if c):
        print(row)
        hits += 1
print(f'GET IT N GO matches: {hits}')

# Also count $100K MONOPOLY DOUBLER entries
hits = 0
for row in ws.iter_rows(values_only=True):
    if row and any('MONOPOLY DOUBLER' in str(c).upper() for c in row if c):
        print(row)
        hits += 1
print(f'MONOPOLY DOUBLER matches: {hits}')
