import requests
HEADERS = {'User-Agent': 'Mozilla/5.0'}
r = requests.get('https://www.molottery.com/monthly-winner/MonthlyWinningsWeb.xlsx',
                 headers=HEADERS, timeout=60)
print('status:', r.status_code, 'len:', len(r.content), 'ct:', r.headers.get('content-type'))
with open('_mo_winners.xlsx', 'wb') as f:
    f.write(r.content)
# Try parsing
try:
    from openpyxl import load_workbook
    wb = load_workbook('_mo_winners.xlsx', read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'--- sheet: {sn} (max_row={ws.max_row}, max_col={ws.max_column}) ---')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            print(row)
            if i >= 12:
                print('... (truncated)')
                break
except Exception as e:
    print('openpyxl failed:', e)
